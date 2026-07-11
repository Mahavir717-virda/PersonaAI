"""Smart Attachment Text Extraction Service.

Routes attachment files to the appropriate extractor based on file type,
returning extracted plaintext for downstream AI summarization.
"""

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

# Size threshold: attachments larger than this are skipped for auto-processing
MAX_AUTO_PROCESS_BYTES = 5 * 1024 * 1024  # 5 MB

# Supported text-extractable MIME types
TEXT_MIME_TYPES = {
    "text/plain", "text/csv", "text/markdown", "text/html",
    "text/x-python", "text/x-java-source", "text/x-c",
}

IMAGE_MIME_TYPES = {
    "image/jpeg", "image/png", "image/gif", "image/webp",
    "image/bmp", "image/tiff",
}


def _get_file_category(filename: str, content_type: str | None) -> str:
    """Determine the processing category from filename extension and MIME type."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    ct = (content_type or "").lower()

    if ext == "pdf" or "pdf" in ct:
        return "pdf"
    if ext in ("docx",) or "wordprocessingml" in ct:
        return "docx"
    if ext in ("xlsx", "xls") or "spreadsheetml" in ct:
        return "xlsx"
    if ext == "csv" or "csv" in ct:
        return "csv"
    if ext in ("txt", "md", "markdown", "log", "py", "js", "ts", "java", "c", "cpp", "h", "json", "yaml", "yml", "xml", "html", "htm", "css", "sql", "sh", "bat", "ini", "cfg", "toml"):
        return "text"
    if ext in ("pptx",) or "presentationml" in ct:
        return "pptx"
    if ext in ("jpg", "jpeg", "png", "gif", "webp", "bmp", "tiff") or ct in IMAGE_MIME_TYPES:
        return "image"

    # Fallback: check if content_type hints at text
    if ct in TEXT_MIME_TYPES or ct.startswith("text/"):
        return "text"

    return "unsupported"


class AttachmentExtractor:
    """Unified text extraction router for email attachments."""

    def extract_text(
        self,
        filename: str,
        content_bytes: bytes,
        content_type: str | None = None,
    ) -> str | None:
        """Extract readable text from an attachment.

        Returns:
            Extracted text string, or None if:
            - The file type requires vision model processing (images, scanned PDFs)
            - The file type is unsupported
        """
        category = _get_file_category(filename, content_type)
        logger.info(
            "[AttachmentExtractor] Processing '%s' (type=%s, category=%s, size=%d bytes)",
            filename, content_type, category, len(content_bytes),
        )

        try:
            if category == "pdf":
                return self._extract_pdf(content_bytes)
            elif category == "docx":
                return self._extract_docx(content_bytes)
            elif category == "text":
                return self._extract_txt(content_bytes)
            elif category == "csv":
                return self._extract_csv(content_bytes)
            elif category == "xlsx":
                return self._extract_xlsx(content_bytes)
            elif category == "image":
                return None  # Signal: requires vision model
            else:
                logger.info("[AttachmentExtractor] Unsupported file category: %s", category)
                return None
        except Exception as e:
            logger.error(
                "[AttachmentExtractor] Extraction failed for '%s': %s",
                filename, str(e), exc_info=True,
            )
            return None

    def get_file_category(self, filename: str, content_type: str | None = None) -> str:
        """Public accessor for file category routing."""
        return _get_file_category(filename, content_type)

    def _extract_pdf(self, content_bytes: bytes) -> str | None:
        """Extract text from PDF using PyMuPDF (fitz).

        If extracted text is very sparse relative to page count,
        the PDF is likely scanned — returns None to trigger vision fallback.
        """
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.error("[AttachmentExtractor] PyMuPDF (fitz) is not installed. Cannot extract PDF text.")
            return None

        doc = fitz.open(stream=content_bytes, filetype="pdf")
        pages_text: list[str] = []

        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text").strip()
            if text:
                pages_text.append(f"--- Page {page_num + 1} ---\n{text}")

        doc.close()

        if not pages_text:
            logger.info("[AttachmentExtractor] PDF has no extractable text (likely scanned). Falling back to vision.")
            return None

        # Scanned PDF heuristic: if average chars per page is very low, treat as scanned
        total_chars = sum(len(t) for t in pages_text)
        page_count = len(doc) if len(doc) > 0 else 1
        avg_chars_per_page = total_chars / page_count

        if avg_chars_per_page < 50 and page_count > 1:
            logger.info(
                "[AttachmentExtractor] PDF appears scanned (avg %d chars/page across %d pages). Falling back to vision.",
                avg_chars_per_page, page_count,
            )
            return None

        full_text = "\n\n".join(pages_text)
        # Truncate to a reasonable size for LLM context (roughly 30k chars ≈ 8k tokens)
        if len(full_text) > 30000:
            full_text = full_text[:30000] + "\n\n[...content truncated for AI processing...]"

        return full_text

    def _extract_docx(self, content_bytes: bytes) -> str | None:
        """Extract text from DOCX using python-docx."""
        try:
            from docx import Document
        except ImportError:
            logger.error("[AttachmentExtractor] python-docx is not installed. Cannot extract DOCX text.")
            return None

        doc = Document(io.BytesIO(content_bytes))
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

        if not paragraphs:
            return None

        full_text = "\n".join(paragraphs)
        if len(full_text) > 30000:
            full_text = full_text[:30000] + "\n\n[...content truncated for AI processing...]"

        return full_text

    def _extract_txt(self, content_bytes: bytes) -> str | None:
        """Decode plain text files with encoding detection fallback."""
        # Try UTF-8 first
        try:
            text = content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            # Fallback: try chardet for encoding detection
            try:
                import chardet
                detected = chardet.detect(content_bytes)
                encoding = detected.get("encoding", "latin-1") or "latin-1"
                text = content_bytes.decode(encoding, errors="replace")
            except Exception:
                text = content_bytes.decode("latin-1", errors="replace")

        text = text.strip()
        if not text:
            return None

        if len(text) > 30000:
            text = text[:30000] + "\n\n[...content truncated for AI processing...]"

        return text

    def _extract_csv(self, content_bytes: bytes) -> str | None:
        """Parse CSV and format as a readable text table (first 100 rows)."""
        try:
            text = content_bytes.decode("utf-8", errors="replace")
        except Exception:
            return None

        reader = csv.reader(io.StringIO(text))
        rows: list[list[str]] = []
        for i, row in enumerate(reader):
            if i >= 100:
                break
            rows.append(row)

        if not rows:
            return None

        # Format as readable table
        output_lines: list[str] = []
        if rows:
            header = rows[0]
            output_lines.append(" | ".join(header))
            output_lines.append("-" * len(output_lines[0]))
            for row in rows[1:]:
                output_lines.append(" | ".join(row))

        result = "\n".join(output_lines)
        total_rows = sum(1 for _ in csv.reader(io.StringIO(text))) - 1
        if total_rows > 100:
            result += f"\n\n[...showing 100 of {total_rows} rows...]"

        return result

    def _extract_xlsx(self, content_bytes: bytes) -> str | None:
        """Extract spreadsheet data from XLSX using openpyxl (first 100 rows per sheet)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("[AttachmentExtractor] openpyxl is not installed. Cannot extract XLSX data.")
            return None

        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
        sheet_texts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 100:
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if not rows:
                continue

            lines = [f"=== Sheet: {sheet_name} ==="]
            # Header
            lines.append(" | ".join(rows[0]))
            lines.append("-" * len(lines[-1]))
            for row in rows[1:]:
                lines.append(" | ".join(row))

            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0
            if total_rows > 100:
                lines.append(f"\n[...showing 100 of {total_rows} rows, {total_cols} columns...]")

            sheet_texts.append("\n".join(lines))

        wb.close()

        if not sheet_texts:
            return None

        result = "\n\n".join(sheet_texts)
        if len(result) > 30000:
            result = result[:30000] + "\n\n[...content truncated for AI processing...]"

        return result
